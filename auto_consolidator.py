import os
import sys
import traceback
from openpyxl import load_workbook, Workbook
import tkinter as tk
from tkinter import filedialog, messagebox, ttk, scrolledtext, font as tkFont
from datetime import datetime
from pathlib import Path
import logging
import threading
from ttkthemes import ThemedTk
from dataclasses import dataclass
from contextlib import contextmanager
from typing import Optional, List, Callable, Union

# --- Lightweight DataFrame Replacement ---
class CellMapData:
    """
    Lightweight replacement for pandas DataFrame specifically for Cell Map data.
    Provides the minimal interface needed by the consolidator application.
    Uses only built-in Python types and openpyxl for maximum efficiency.
    """
    
    def __init__(self, data: List[List], columns: List[str]):
        """
        Initialize with data and column names.
        
        Args:
            data: List of rows, where each row is a list of values
            columns: List of column names
        """
        if not columns:
            raise ValueError("Columns cannot be empty")
        if data and len(data[0]) != len(columns):
            raise ValueError("Number of columns must match data width")
            
        self._data = data
        self._columns = columns
        self._column_indices = {col: idx for idx, col in enumerate(columns)}
    
    @property
    def columns(self) -> List[str]:
        """Return list of column names"""
        return self._columns.copy()
    
    def __len__(self) -> int:
        """Return number of rows"""
        return len(self._data)
    
    def __getitem__(self, column_name: str) -> List:
        """Get column data by name"""
        if column_name not in self._column_indices:
            raise KeyError(f"Column '{column_name}' not found")
        col_idx = self._column_indices[column_name]
        return [row[col_idx] for row in self._data]
    
    def iterrows(self):
        """
        Iterate over rows, yielding (index, row_data) tuples.
        row_data is a dict-like object that supports column access.
        """
        for idx, row in enumerate(self._data):
            row_dict = CellMapRow(dict(zip(self._columns, row)))
            yield idx, row_dict
    
    def isnull(self):
        """Return a CellMapData with boolean values indicating null/empty values"""
        null_data = []
        for row in self._data:
            null_row = [val is None or (isinstance(val, str) and val.strip() == "") for val in row]
            null_data.append(null_row)
        return CellMapData(null_data, self._columns)
    
    def duplicated(self, subset: Optional[List[str]] = None):
        """Return a CellMapData with boolean values indicating duplicate rows"""
        if subset is None:
            subset = self._columns
        
        # Get indices for subset columns
        subset_indices = [self._column_indices[col] for col in subset if col in self._column_indices]
        
        seen_rows = set()
        duplicate_flags = []
        
        for row in self._data:
            # Create tuple of subset values for comparison
            subset_tuple = tuple(row[idx] for idx in subset_indices)
            is_duplicate = subset_tuple in seen_rows
            duplicate_flags.append([is_duplicate])  # Single column result
            seen_rows.add(subset_tuple)
        
        return CellMapData(duplicate_flags, ["is_duplicate"])

class CellMapRow:
    """
    Dictionary-like object to represent a row from CellMapData.
    Supports both dict-style and attribute-style access.
    """
    
    def __init__(self, data: dict):
        self._data = data
    
    def __getitem__(self, key: str):
        """Dictionary-style access: row["column_name"]"""
        return self._data[key]
    
    def __contains__(self, key: str) -> bool:
        """Support 'in' operator"""
        return key in self._data
    
    def get(self, key: str, default=None):
        """Dict-style get method"""
        return self._data.get(key, default)

def read_excel_to_cellmapdata(file_path: str, sheet_name: Optional[str] = None) -> CellMapData:
    """
    Read an Excel file using openpyxl and return a CellMapData instance.
    
    Args:
        file_path: Path to the Excel file
        sheet_name: Name of the sheet to read (if None, uses active sheet)
    
    Returns:
        CellMapData instance containing the Excel data
    
    Raises:
        ValueError: If the file cannot be read or has no data
    """
    try:
        # Open the workbook
        wb = load_workbook(file_path, read_only=True, data_only=True)
          # Select the worksheet
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found in workbook")
            ws = wb[sheet_name]
        else:
            ws = wb.active
        
        # Ensure we have a valid worksheet
        if ws is None:
            raise ValueError("No active worksheet found in workbook")
        
        # Read all data from the worksheet
        data = []
        columns = []
        
        # Get the data range (skip empty rows/columns)
        if ws.max_row == 1 and ws.max_column == 1:
            # Check if the single cell is empty
            cell_value = ws.cell(1, 1).value
            if cell_value is None or str(cell_value).strip() == "":
                raise ValueError("Excel sheet appears to be empty")
        
        # Read header row (first row)
        first_row = True
        for row in ws.iter_rows(values_only=True):
            # Skip completely empty rows
            if not any(cell is not None and str(cell).strip() for cell in row):
                continue
                
            if first_row:
                # First non-empty row becomes the header
                columns = [str(cell).strip() if cell is not None else f"Column_{i}" 
                          for i, cell in enumerate(row)]
                first_row = False
            else:
                # Convert None values to empty strings and ensure consistent row length
                row_data = []
                for i, cell in enumerate(row):
                    if i >= len(columns):
                        break  # Don't exceed column count
                    if cell is None:
                        row_data.append("")
                    else:
                        row_data.append(str(cell).strip() if isinstance(cell, str) else cell)
                
                # Pad row to match column count if necessary
                while len(row_data) < len(columns):
                    row_data.append("")
                
                data.append(row_data)
        
        wb.close()
        
        if not columns:
            raise ValueError("No valid header row found in Excel file")
        
        return CellMapData(data, columns)
        
    except Exception as e:
        raise ValueError(f"Failed to read Excel file '{file_path}': {str(e)}")

# --- Configuration and Setup ---
@dataclass
class ConsolidatorConfig:
    """Configuration class for the Excel Consolidator"""
    cell_map_path: str = ""
    consolidation_path: str = ""
    consolidation_sheet: str = "General Consolidation"
    header_row: int = 4
    data_start_row: int = 5
    clear_existing_data: bool = True
    max_file_size_mb: int = 50

class Constants:
    """Application constants"""
    DEFAULT_SHEET_NAME = "General Consolidation"
    DEFAULT_HEADER_ROW = 4
    DEFAULT_DATA_START_ROW = 5
    CELL_MAP_FILENAME = "Cell Map.xlsx"
    MAX_EMPTY_ROWS_BEFORE_STOP = 20
    MAX_FILE_SIZE_MB = 50
    PROGRESS_UPDATE_INTERVAL = 100
    MIN_ROW_NUMBER = 1
    MAX_ROW_NUMBER = 1000
    LOG_DATE_FORMAT = "%Y-%m-%d %H:%M:%S"

# Context manager for safe workbook handling
@contextmanager
def open_workbook(file_path, **kwargs):
    """Context manager for safe workbook operations"""
    wb = None
    try:
        wb = load_workbook(file_path, **kwargs)
        yield wb
    finally:
        if wb and hasattr(wb, 'close'):
            try:
                wb.close()
            except:
                pass  # Ignore close errors

LOG_FILE = "consolidator_errors.log"
logging.basicConfig(
    filename=LOG_FILE,
    level=logging.ERROR,
    format="%(asctime)s - %(levelname)s - %(message)s\n%(traceback)s\n",
)

# --- Core Business Logic ---
class ExcelConsolidator:
    def __init__(self, config: ConsolidatorConfig, logger: Optional[logging.Logger] = None):
        self.config = config
        self.logger = logger or logging.getLogger(__name__)
        self.cell_map_df = None
          # Validate and resolve file paths
        self.cell_map_path = self._resolve_file_path(config.cell_map_path, "Cell Map")
        self.consolidation_path = self._resolve_file_path(config.consolidation_path, "Consolidation")
        
        # Validate sheet name
        self.consolidation_sheet = ValidationUtils.validate_sheet_name(config.consolidation_sheet)

    def _resolve_file_path(self, path_str: str, file_type: str) -> Path:
        """Safely resolve file paths with proper error handling and security checks"""
        path = FileHandler.validate_file_path(path_str, file_type)
        
        # Check file size
        if not FileHandler.check_file_size(path, self.config.max_file_size_mb):
            self.logger.warning(f"Large {file_type} file detected: {path.name}")
            
        return path

    def _validate_row_input(self, value: int, field_name: str) -> int:
        """Comprehensive row number validation"""
        if not isinstance(value, int) or not Constants.MIN_ROW_NUMBER <= value <= Constants.MAX_ROW_NUMBER:
            raise ValueError(f"{field_name} must be between {Constants.MIN_ROW_NUMBER} and {Constants.MAX_ROW_NUMBER}")
        return value

    def _validate_cell_map(self):
        """
        Validates the Cell Map file structure and content.
        Raises ValueError on failure.
        """
        required_columns = [
            "Source Sheet",
            "Source Cell",
            "Destination Column (Consolidation)",
        ]

        try:
            df = read_excel_to_cellmapdata(str(self.cell_map_path))
        except Exception as e:
            raise ValueError(f"Failed to read Cell Map file: {e}")

        missing_cols = [col for col in required_columns if col not in df.columns]
        if missing_cols:
            raise ValueError(f"Cell Map is missing required columns: {', '.join(missing_cols)}")

        # Check for null/empty values using our CellMapData interface
        null_check = df.isnull()
        has_nulls = any(any(row) for row in null_check._data)
        if has_nulls:
            raise ValueError("Cell Map contains empty cells. Please fill all values.")
        
        # Check for duplicates using our CellMapData interface
        dup_check = df.duplicated(subset=required_columns)
        has_duplicates = any(row[0] for row in dup_check._data)
        if has_duplicates:
            raise ValueError("Cell Map contains duplicate mappings. Please remove them.")

        self.cell_map_df = df
        self.logger.info(f"Cell Map validated successfully: {len(df)} mappings loaded")
        return True

    def _validate_consolidation_file(self):
        """
        Validates the consolidation file and required sheet/columns.
        Uses the defined header row.
        Raises ValueError on failure.
        """
        if self.cell_map_df is None:
            raise ValueError("Cell Map must be validated first before validating consolidation file")
            
        # Validate row configuration
        header_row = self._validate_row_input(self.config.header_row, "Header row")
        data_start_row = self._validate_row_input(self.config.data_start_row, "Data start row")
        
        if data_start_row <= header_row:
            raise ValueError("Data start row must be after the header row")

        try:
            with open_workbook(self.consolidation_path, read_only=True) as wb:
                if self.config.consolidation_sheet not in wb.sheetnames:
                    available_sheets = ", ".join(wb.sheetnames)
                    raise ValueError(f"Sheet '{self.config.consolidation_sheet}' not found. Available sheets: [{available_sheets}]")

                ws = wb[self.config.consolidation_sheet]
                
                # Check if we have enough rows
                if ws.max_row < header_row:
                    raise ValueError(f"Consolidation sheet has less than {header_row} rows. Cannot find header row.")
                
                # Fetch the header from the defined header row
                header_row_data = [cell.value for cell in ws[header_row]]
                dest_columns = set(self.cell_map_df["Destination Column (Consolidation)"])
                
                # Clean and normalize header data
                header_row_data = [str(h).strip() if h is not None else "" for h in header_row_data]
                dest_columns = {str(d).strip() for d in dest_columns}

                missing_cols = dest_columns - set(header_row_data)
                if missing_cols:
                    self.logger.error(f"Header data from row {header_row}: {header_row_data}")
                    self.logger.error(f"Required destination columns: {dest_columns}")
                    raise ValueError(f"Missing destination columns in consolidation sheet (row {header_row}): {', '.join(sorted(list(missing_cols)))}")
                    
        except Exception as e:
            if isinstance(e, ValueError):
                raise
            raise ValueError(f"Cannot validate consolidation file: {e}")

        self.logger.info(f"Consolidation file validated successfully")
        return True

    def _validate_estimate_file(self, file_path):
        """
        Validates a single estimate file contains the sheets and cells defined in the cell map.
        Raises ValueError on failure.
        """
        if self.cell_map_df is None:
            raise ValueError("Cell Map must be validated first before validating estimate files")
            
        try:
            current_file_path = self._resolve_file_path(file_path, "Estimate")
            
            with open_workbook(current_file_path, read_only=True) as wb:
                for _, row in self.cell_map_df.iterrows():
                    sheet_name = row["Source Sheet"]
                    cell_ref = row["Source Cell"]
                    
                    if sheet_name not in wb.sheetnames:
                        available_sheets = ", ".join(wb.sheetnames)
                        raise ValueError(f"In file '{current_file_path.name}', required sheet '{sheet_name}' not found. Available sheets: [{available_sheets}]")

                    ws = wb[sheet_name]
                    
                    try:
                        _ = ws[cell_ref]
                    except Exception as e:
                        raise ValueError(f"In file '{current_file_path.name}', cell '{cell_ref}' could not be accessed in sheet '{sheet_name}'. Error: {e}")
                        
        except ValueError:
            raise
        except Exception as e:
            raise ValueError(f"Cannot validate estimate file '{Path(file_path).name}': {e}")
            
        self.logger.info(f"Estimate file validated: {Path(file_path).name}")
        return True

    def run_consolidation(self, estimate_files: List[str], progress_callback: Optional[Callable] = None):
        """
        Performs the consolidation, updating progress via the callback.
        Enhanced with better error handling and progress tracking.
        """
        if self.cell_map_df is None:
            raise ValueError("Cell Map must be validated first before running consolidation")
            
        try:
            # Load consolidation workbook and sheet
            cons_wb = load_workbook(self.consolidation_path)
            if self.config.consolidation_sheet not in cons_wb.sheetnames:
                raise ValueError(f"Critical: Consolidation sheet '{self.config.consolidation_sheet}' not found in workbook '{self.consolidation_path.name}'")
            
            cons_ws = cons_wb[self.config.consolidation_sheet]

            # Validate row configuration
            header_row = self._validate_row_input(self.config.header_row, "Header row")
            data_start_row = self._validate_row_input(self.config.data_start_row, "Data start row")

            # Find header row and column indices
            if cons_ws.max_row < header_row:
                raise ValueError(f"Consolidation sheet '{self.config.consolidation_sheet}' has less than {header_row} rows. Cannot find header row for writing.")

            header_values = [cell.value for cell in cons_ws[header_row]]
            col_indices = {name: idx + 1 for idx, name in enumerate(header_values) if name is not None}

            # Clear existing data if requested
            if self.config.clear_existing_data:
                self._clear_existing_data(cons_ws, col_indices, data_start_row, progress_callback, len(estimate_files))            # Process estimate files
            current_data_row = data_start_row
            item_number = 1  # Start item numbering from 1
            for i, est_file_path_str in enumerate(estimate_files):
                est_file_full_path = Path(est_file_path_str).resolve()
                
                # Set filename in first column (Column A)
                cons_ws.cell(row=current_data_row, column=1, value=est_file_full_path.stem)
                
                # Set item number in second column (Column B) 
                cons_ws.cell(row=current_data_row, column=2, value=item_number)
                
                # Process each mapping with enhanced validation
                for _, map_row in self.cell_map_df.iterrows():
                    src_sheet = ValidationUtils.validate_sheet_name(map_row["Source Sheet"])
                    src_cell = str(map_row["Source Cell"]).strip()
                    dest_col_name = str(map_row["Destination Column (Consolidation)"]).strip()

                    if dest_col_name not in col_indices:
                        self.logger.warning(f"Destination column '{dest_col_name}' not found in consolidation header. Skipping.")
                        continue

                    dest_col_idx = col_indices[dest_col_name]
                    
                    # Create Excel formula for linking
                    formula = f"='{est_file_full_path.parent}\\[{est_file_full_path.name}]{src_sheet}'!{src_cell}"
                    
                    # Write the formula directly (no validation needed for our generated formulas)
                    cons_ws.cell(row=current_data_row, column=dest_col_idx, value=formula)
                
                current_data_row += 1
                item_number += 1  # Increment item number for next row

                if progress_callback:
                    progress_callback("processing", i + 1, len(estimate_files), f"Processed {est_file_full_path.name} (Item #{item_number-1})")

            # Save output file
            if progress_callback:
                progress_callback("saving", 1, 1, "Saving consolidated file...")
                
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = self.consolidation_path.parent / f"Consolidation_AutoLinked_{timestamp}.xlsx"
            cons_wb.save(output_file)
            cons_wb.close()
            
            self.logger.info(f"Consolidation completed: {output_file}")
            return output_file

        except Exception as e:
            self.logger.error("Error during consolidation", exc_info=True)
            raise

    def _clear_existing_data(self, cons_ws, col_indices: dict, data_start_row: int, progress_callback: Optional[Callable], total_files: int):
        """
        Clear existing data from the consolidation sheet
        """
        if self.cell_map_df is None:
            raise ValueError("Cell Map must be validated before clearing data")
            
        if progress_callback:
            progress_callback("clearing", 0, total_files, "Clearing existing data...")
        
        # Determine columns to clear (only those that are mapped)
        columns_to_clear_indices = [col_indices[dest_col] for dest_col in self.cell_map_df["Destination Column (Consolidation)"] if dest_col in col_indices]
        
        # Clear data with safety limits
        rows_cleared_consecutively_empty = 0
        max_rows_to_scan = cons_ws.max_row
        
        for r_idx in range(data_start_row, max_rows_to_scan + 1):            # Check if the first cell has content
            first_cell_val = cons_ws.cell(row=r_idx, column=1).value
            if first_cell_val is not None and str(first_cell_val).strip() != "":
                # Clear mapped columns and filename column
                for col_idx_to_clear in columns_to_clear_indices:
                    cons_ws.cell(row=r_idx, column=col_idx_to_clear, value=None)
                # Clear filename column (Column A) and item number column (Column B)
                cons_ws.cell(row=r_idx, column=1, value=None)
                cons_ws.cell(row=r_idx, column=2, value=None)
                rows_cleared_consecutively_empty = 0
            else:
                rows_cleared_consecutively_empty += 1
            
            if rows_cleared_consecutively_empty > Constants.MAX_EMPTY_ROWS_BEFORE_STOP:
                break
        
        if progress_callback:
            progress_callback("clearing", total_files, total_files, "Clearing complete.")


# --- Graphical User Interface ---
class ConsolidatorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Auto Consolidator")
        self.root.configure(bg="#e0e0e0") # Light gray background for the main window

        # --- Style Configuration ---
        self.style = ttk.Style(self.root)
        try:
            # Try a theme that often looks more modern
            self.style.theme_use('clam')
        except tk.TclError:
            print("Clam theme not available, using default.")

        # Define custom fonts
        self.default_font = tkFont.Font(family="Segoe UI", size=10)
        self.label_font = tkFont.Font(family="Segoe UI", size=10, weight="bold")
        self.button_font = tkFont.Font(family="Segoe UI", size=10, weight="bold")

        # Configure default styles for ttk widgets
        self.style.configure("TLabel", font=self.default_font, background="#e0e0e0", foreground="#333333")
        self.style.configure("TButton", font=self.button_font, padding=6)
        self.style.configure("TEntry", font=self.default_font, padding=5)
        self.style.configure("TListbox", font=self.default_font) # Listbox is a tk widget, but let's try
        self.style.configure("TProgressbar", thickness=20)

        # Custom style for the main action button
        self.style.configure("Accent.TButton", font=self.button_font, background="#0078D4", foreground="white")
        self.style.map("Accent.TButton",
            background=[('active', '#005A9E')] # Darker blue on hover/press
        )
        # Custom style for secondary buttons
        self.style.configure("Secondary.TButton", font=self.default_font)
        self.style.configure("Success.TButton", font=self.default_font, background="#28a745", foreground="white") # Green
        self.style.map("Success.TButton", background=[('active', '#1e7e34')])
        
        self.cell_map_path = tk.StringVar()
        self.consolidation_path = tk.StringVar()
        self.consolidation_sheet = tk.StringVar(value=Constants.DEFAULT_SHEET_NAME)
        # --- Feature: Configurable Header/Data Rows ---
        self.header_row_var = tk.StringVar(value=str(Constants.DEFAULT_HEADER_ROW))
        self.data_start_row_var = tk.StringVar(value=str(Constants.DEFAULT_DATA_START_ROW))
        # --- Feature: Clear Existing Data ---
        self.clear_data_var = tk.BooleanVar(value=True) # Default to True for a clean slate
        
        self.estimate_files = []
        self.output_file_path = None # For "Open Output" feature        # --- Improvement 1: Auto-populate Cell Map Path ---
        # Attempt to find Cell Map.xlsx in the application's directory
        default_cell_map_name = Constants.CELL_MAP_FILENAME
        # Try current working directory first
        potential_cell_map_path = Path(os.getcwd()) / default_cell_map_name
        if not potential_cell_map_path.exists():
            # Try script's directory if different and CWD doesn't have it
            try:
                script_dir = Path(__file__).resolve().parent
                potential_cell_map_path = script_dir / default_cell_map_name
            except NameError: # __file__ is not defined (e.g. in interactive interpreter)
                pass # Stick with CWD path or empty

        if potential_cell_map_path.exists():
            self.cell_map_path.set(str(potential_cell_map_path))
        else:
            self.cell_map_path.set("") # Leave blank if not found

        self._setup_gui()

    def _setup_gui(self):
        # Main frame with padding
        main_frame = ttk.Frame(self.root, padding="20 20 20 20", style="TFrame")
        main_frame.pack(expand=True, fill=tk.BOTH)
        self.style.configure("TFrame", background="#e0e0e0")


        # --- File Selection Group ---
        files_group = ttk.Labelframe(main_frame, text="Configuration", padding="10 10 10 10", style="TLabelframe")
        files_group.grid(row=0, column=0, padx=10, pady=10, sticky="ew")
        self.style.configure("TLabelframe", font=self.label_font, background="#e0e0e0", foreground="#333333")
        self.style.configure("TLabelframe.Label", font=self.label_font, background="#e0e0e0", foreground="#333333")


        row_idx = 0
        ttk.Label(files_group, text="Cell Map File:").grid(row=row_idx, column=0, sticky="e", padx=5, pady=5)
        ttk.Entry(files_group, textvariable=self.cell_map_path, width=50).grid(row=row_idx, column=1, padx=5, pady=5)
        ttk.Button(files_group, text="Browse", command=self._select_cell_map, style="Secondary.TButton").grid(row=row_idx, column=2, padx=5, pady=5)
        row_idx += 1
        ttk.Label(files_group, text="Consolidation File:").grid(row=row_idx, column=0, sticky="e", padx=5, pady=5)
        ttk.Entry(files_group, textvariable=self.consolidation_path, width=50).grid(row=row_idx, column=1, padx=5, pady=5)
        ttk.Button(files_group, text="Browse", command=self._select_consolidation_file, style="Secondary.TButton").grid(row=row_idx, column=2, padx=5, pady=5)
        row_idx += 1
        ttk.Label(files_group, text="Consolidation Sheet:").grid(row=row_idx, column=0, sticky="e", padx=5, pady=5)
        ttk.Entry(files_group, textvariable=self.consolidation_sheet, width=30).grid(row=row_idx, column=1, sticky="w", padx=5, pady=5)
        row_idx += 1
        # --- Feature: Configurable Header/Data Rows GUI ---
        ttk.Label(files_group, text="Header Row (Consolidation):").grid(row=row_idx, column=0, sticky="e", padx=5, pady=5)
        ttk.Entry(files_group, textvariable=self.header_row_var, width=5).grid(row=row_idx, column=1, sticky="w", padx=5, pady=5)
        row_idx += 1
        ttk.Label(files_group, text="Data Start Row (Consolidation):").grid(row=row_idx, column=0, sticky="e", padx=5, pady=5)
        ttk.Entry(files_group, textvariable=self.data_start_row_var, width=5).grid(row=row_idx, column=1, sticky="w", padx=5, pady=5)
        
        files_group.columnconfigure(1, weight=1) # Make entry column expandable

        # --- Estimate Files Group ---
        estimates_group = ttk.Labelframe(main_frame, text="Estimate Files", padding="10 10 10 10", style="TLabelframe")
        estimates_group.grid(row=1, column=0, padx=10, pady=10, sticky="ew")

        # Note: Listbox is a standard tk widget, not ttk. Styling is more limited.
        # For a fully themed list, you might need a ttk.Treeview or a custom widget.
        self.estimate_listbox = tk.Listbox(estimates_group, selectmode=tk.MULTIPLE, width=60, height=6,
                                           font=self.default_font, bg="white", fg="#333333",
                                           selectbackground="#0078D4", selectforeground="white",
                                           borderwidth=1, relief="solid")
        self.estimate_listbox.grid(row=0, column=0, columnspan=2, sticky="ewns", padx=5, pady=5)
        estimates_group.rowconfigure(0, weight=1)
        estimates_group.columnconfigure(0, weight=1)
        
        buttons_frame_estimates = ttk.Frame(estimates_group, style="TFrame")
        buttons_frame_estimates.grid(row=0, column=2, sticky="ns", padx=5, pady=5)
        ttk.Button(buttons_frame_estimates, text="Add Files", command=self._select_estimate_files, style="Secondary.TButton").pack(fill=tk.X, pady=2)
        ttk.Button(buttons_frame_estimates, text="Clear", command=self._clear_estimate_files, style="Secondary.TButton").pack(fill=tk.X, pady=2)


        # --- Progress and Log Group ---
        status_group = ttk.Labelframe(main_frame, text="Status & Actions", padding="10 10 10 10", style="TLabelframe")
        status_group.grid(row=2, column=0, padx=10, pady=10, sticky="ew")

        self.progress = ttk.Progressbar(status_group, orient="horizontal", length=400, mode="determinate")
        self.progress.pack(fill=tk.X, padx=5, pady=5) # Use pack here for simplicity within this group

        self.log_text = scrolledtext.ScrolledText(status_group, width=70, height=8, state="disabled",
                                                  font=tkFont.Font(family="Consolas", size=9), # Monospaced for logs
                                                  bg="#f0f0f0", fg="#333333", relief="solid", borderwidth=1)
        self.log_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        # Tag configurations for log_text would remain in _log_message

        action_buttons_frame = ttk.Frame(main_frame, style="TFrame")
        action_buttons_frame.grid(row=3, column=0, padx=10, pady=10, sticky="ew")
        
        # --- Feature: Clear Existing Data Checkbox ---
        self.clear_data_checkbox = ttk.Checkbutton(action_buttons_frame, text="Clear existing data in target rows before processing", variable=self.clear_data_var)
        self.clear_data_checkbox.pack(side=tk.LEFT, padx=(0, 20), pady=5)

        self.run_button = ttk.Button(action_buttons_frame, text="Run Consolidation", command=self._start_consolidation_thread, style="Accent.TButton", state=tk.DISABLED)
        self.run_button.pack(side=tk.LEFT, padx=5, pady=5, fill=tk.X, expand=True)
        
        # --- Feature: Open Output Buttons ---
        self.open_output_button = ttk.Button(action_buttons_frame, text="Open Output File", command=self._open_output_file, style="Success.TButton", state=tk.DISABLED)
        self.open_output_button.pack(side=tk.LEFT, padx=5, pady=5)
        self.open_folder_button = ttk.Button(action_buttons_frame, text="Open Output Folder", command=self._open_output_folder, style="Success.TButton", state=tk.DISABLED)
        self.open_folder_button.pack(side=tk.LEFT, padx=5, pady=5)
        
        main_frame.columnconfigure(0, weight=1)
        self._update_run_button_state()

    def _select_file(self, title, string_var): # Unchanged
        file = filedialog.askopenfilename(title=title, filetypes=[("Excel Files", "*.xlsx")])
        if file: string_var.set(file)
    def _select_cell_map(self): self._select_file("Select Cell Map File", self.cell_map_path) # Unchanged
    def _select_consolidation_file(self): self._select_file("Select Consolidation File", self.consolidation_path) # Unchanged
    def _select_estimate_files(self): # Unchanged
        files = filedialog.askopenfilenames(title="Select Estimate Files", filetypes=[("Excel Files", "*.xlsx")])
        if files:
            self.estimate_files.extend([f for f in files if f not in self.estimate_files])
            self._update_listbox()
    def _update_listbox(self): # Unchanged
        self.estimate_listbox.delete(0, tk.END)
        for f in self.estimate_files: self.estimate_listbox.insert(tk.END, f)
        self._update_run_button_state()
    def _clear_estimate_files(self): # Unchanged
        self.estimate_files = []
        self._update_listbox()
    def _update_run_button_state(self): # Unchanged
        if self.estimate_files: self.run_button.config(state=tk.NORMAL)
        else: self.run_button.config(state=tk.DISABLED)

    def _log_message(self, msg, level="INFO"): # Unchanged
        self.log_text.config(state="normal")
        self.log_text.insert(tk.END, f"[{level}] {msg}\n", level)
        self.log_text.config(state="disabled")
        self.log_text.see(tk.END)
        self.root.update_idletasks()
        self.log_text.tag_config("INFO", foreground="black")
        self.log_text.tag_config("SUCCESS", foreground="green")
        self.log_text.tag_config("WARN", foreground="orange")
        self.log_text.tag_config("ERROR", foreground="red")

    def _progress_callback(self, phase: str, current: int, total: int, message: str):
        """Enhanced progress tracking with phase information"""
        phase_weights = {"validation": 0.1, "clearing": 0.1, "processing": 0.7, "saving": 0.1}
        
        # Calculate base progress from completed phases
        completed_phases = []
        if phase in ["clearing", "processing", "saving"]:
            completed_phases.append("validation")
        if phase in ["processing", "saving"]:
            completed_phases.append("clearing")
        if phase == "saving":
            completed_phases.append("processing")
            
        base_progress = sum(phase_weights.get(p, 0) for p in completed_phases) * 100
        
        # Calculate current phase progress
        phase_progress = 0
        if total > 0:
            phase_progress = (current / total) * phase_weights.get(phase, 0.1) * 100
            
        self.progress['value'] = base_progress + phase_progress
        self._log_message(f"[{phase.upper()}] {message}", "INFO")
        self.root.update_idletasks()

    # Backward compatibility method for simple progress updates
    def _simple_progress_callback(self, current: int, total: int, message: str):
        """Simple progress callback for backward compatibility"""
        if total > 0:
            self.progress['value'] = (current / total) * 100
        else:
            self.progress['value'] = 0
        self._log_message(message, "INFO")
        self.root.update_idletasks()

    # --- Feature: Open Output File/Folder Methods ---
    def _open_output_file(self):
        if self.output_file_path and Path(self.output_file_path).exists():
            try:
                os.startfile(self.output_file_path)
                self._log_message(f"Opened output file: {self.output_file_path}", "INFO")
            except Exception as e:
                self._log_message(f"Failed to open output file: {e}", "ERROR")
                messagebox.showerror("Error", f"Could not open file: {e}")
        else:
            self._log_message("Output file not available or does not exist.", "WARN")
            messagebox.showwarning("Not Found", "Output file is not available or no longer exists.")

    def _open_output_folder(self):
        if self.output_file_path and Path(self.output_file_path).exists():
            try:
                os.startfile(Path(self.output_file_path).parent)
                self._log_message(f"Opened output folder: {Path(self.output_file_path).parent}", "INFO")
            except Exception as e:
                self._log_message(f"Failed to open output folder: {e}", "ERROR")
                messagebox.showerror("Error", f"Could not open folder: {e}")
        else:
            self._log_message("Output folder not available (no output file generated yet).", "WARN")
            messagebox.showwarning("Not Found", "Output folder is not available.")


    def _start_consolidation_thread(self):
        if not self.cell_map_path.get() or not self.consolidation_path.get():
            messagebox.showerror("Missing Info", "Please select the Cell Map and Consolidation files.")
            return
        if not self.estimate_files:
            messagebox.showerror("Missing Info", "Please add at least one estimate spreadsheet.")
            return
        
        # Enhanced validation for sheet name
        try:
            sheet_name = ValidationUtils.validate_sheet_name(self.consolidation_sheet.get())
        except ValueError as e:
            messagebox.showerror("Invalid Sheet Name", str(e))
            return
          
        # --- Feature: Validate Header/Data Start Row Inputs ---
        try:
            header_row = int(self.header_row_var.get())
            data_start_row = int(self.data_start_row_var.get())
            if not Constants.MIN_ROW_NUMBER <= header_row <= Constants.MAX_ROW_NUMBER:
                raise ValueError(f"Header row must be between {Constants.MIN_ROW_NUMBER} and {Constants.MAX_ROW_NUMBER}")
            if not Constants.MIN_ROW_NUMBER <= data_start_row <= Constants.MAX_ROW_NUMBER:
                raise ValueError(f"Data start row must be between {Constants.MIN_ROW_NUMBER} and {Constants.MAX_ROW_NUMBER}")
            if data_start_row <= header_row:
                raise ValueError("Data start row must be after the header row.")
        except ValueError as e:
            messagebox.showerror("Invalid Row Input", f"Please enter valid row numbers. {e}")
            return

        self.run_button.config(state="disabled")
        self.open_output_button.config(state="disabled") # Disable during run
        self.open_folder_button.config(state="disabled") # Disable during run
        self.output_file_path = None # Reset previous output path

        thread = threading.Thread(target=self._run_consolidation_logic, daemon=True)
        thread.start()

    def _run_consolidation_logic(self):
        try:
            self._log_message("Starting validation...", "INFO")
            
            # Create configuration object
            config = ConsolidatorConfig(
                cell_map_path=self.cell_map_path.get(),
                consolidation_path=self.consolidation_path.get(),
                consolidation_sheet=self.consolidation_sheet.get(),
                header_row=int(self.header_row_var.get()),
                data_start_row=int(self.data_start_row_var.get()),
                clear_existing_data=self.clear_data_var.get(),
                max_file_size_mb=Constants.MAX_FILE_SIZE_MB
            )

            consolidator = ExcelConsolidator(config)
            
            self._log_message("Validating Cell Map...", "INFO")
            consolidator._validate_cell_map()
            if consolidator.cell_map_df is None:
                raise ValueError("Critical Internal Error: Cell Map DataFrame not loaded.")
            self._log_message("Cell Map DataFrame loaded.", "SUCCESS")

            self._log_message("Validating consolidation file...", "INFO")
            consolidator._validate_consolidation_file()
            self._log_message("Consolidation file validated.", "SUCCESS")

            if not self.estimate_files:
                raise ValueError("No estimate files selected.")
            
            self._log_message(f"Validating {len(self.estimate_files)} estimate file(s)...", "INFO")
            for f_idx, f_path in enumerate(self.estimate_files):
                self._log_message(f"Validating estimate file {f_idx+1}: {Path(f_path).name}", "INFO")
                consolidator._validate_estimate_file(f_path)
            self._log_message("All estimate files validated.", "SUCCESS")

            self.progress["value"] = 0
            self._log_message("Starting consolidation process...", "INFO")
            
            output_file = consolidator.run_consolidation(self.estimate_files, self._progress_callback)
            
            self.output_file_path = output_file
            success_msg = f"Consolidation complete! Output saved as: {self.output_file_path}"
            self._log_message(success_msg, "SUCCESS")
            messagebox.showinfo("Success", success_msg)
            # Enable open buttons on success
            self.open_output_button.config(state=tk.NORMAL)
            self.open_folder_button.config(state=tk.NORMAL)

        except ValueError as ve:
            logging.error(f"Validation or Logic Error: {str(ve)}", exc_info=False)
            self._log_message(f"Error: {str(ve)}", "ERROR")
            messagebox.showerror("Error", str(ve))
        except Exception as e:
            critical_error_msg = f"A critical error occurred: {type(e).__name__} - {str(e)}"
            logging.error(critical_error_msg, exc_info=True)
            self._log_message(critical_error_msg, "ERROR")
            messagebox.showerror("Critical Error", critical_error_msg)
        finally:
            # Re-enable run button regardless of outcome
            self.root.after(0, lambda: self.run_button.config(state=tk.NORMAL if self.estimate_files else tk.DISABLED))


# --- Utility Classes ---
class FileHandler:
    """Utility class for safe file operations"""
    
    @staticmethod
    def validate_file_path(path_str: str, file_type: str = "File") -> Path:
        """Validate and sanitize file paths to prevent path traversal attacks"""
        if not path_str or not path_str.strip():
            raise ValueError(f"{file_type} path cannot be empty")
            
        # Basic sanitization - remove dangerous characters
        sanitized_path = path_str.strip().replace('..', '').replace('\\\\', '\\')
        
        try:
            path = Path(sanitized_path).resolve()
            
            # Ensure the file exists and is readable
            if not path.exists():
                raise FileNotFoundError(f"{file_type} not found: {path}")
                
            if not path.is_file():
                raise ValueError(f"Path is not a file: {path}")
                
            # Basic security check - ensure it's an Excel file
            if not path.suffix.lower() in ['.xlsx', '.xlsm', '.xls']:
                raise ValueError(f"Invalid file type. Expected Excel file, got: {path.suffix}")
                
            return path
            
        except Exception as e:
            if isinstance(e, (ValueError, FileNotFoundError)):
                raise
            raise ValueError(f"Invalid {file_type} path '{path_str}': {e}")

    @staticmethod
    def check_file_size(file_path: Path, max_size_mb: int = 50) -> bool:
        """Check if file size is within acceptable limits"""
        try:
            file_size_mb = file_path.stat().st_size / (1024 * 1024)
            if file_size_mb > max_size_mb:
                logging.warning(f"Large file detected: {file_path.name} ({file_size_mb:.1f}MB)")
                return False
            return True
        except Exception:
            return True  # If we can't check size, assume it's okay

# Enhanced validation utilities
class ValidationUtils:
    """Utility methods for input validation"""
    
    @staticmethod
    def validate_excel_formula_injection(value: str, allow_formulas: bool = False) -> str:
        """Basic protection against Excel formula injection
        
        Args:
            value: The value to validate
            allow_formulas: If True, allows legitimate Excel formulas starting with =
        """
        if not isinstance(value, str):
            return value
            
        # If formulas are allowed, don't sanitize legitimate Excel formulas
        if allow_formulas:
            return value
            
        # For user input data, sanitize potential formula injection
        if value.strip().startswith(('=', '+', '-', '@')):
            # Log potential injection attempt
            logging.warning(f"Potential formula injection detected in user data: {value[:50]}")
            # Sanitize by prefixing with single quote
            return "'" + value
        return value
    
    @staticmethod
    def validate_sheet_name(sheet_name: str) -> str:
        """Validate Excel sheet names"""
        if not sheet_name or not sheet_name.strip():
            raise ValueError("Sheet name cannot be empty")
            
        # Excel sheet name restrictions
        invalid_chars = ['\\', '/', '*', '[', ']', ':', '?']
        for char in invalid_chars:
            if char in sheet_name:
                raise ValueError(f"Sheet name contains invalid character '{char}'")
                
        if len(sheet_name) > 31:
            raise ValueError("Sheet name cannot exceed 31 characters")
            
        return sheet_name.strip()

class MemoryOptimizer:
    """Utility class for memory management and optimization"""
    
    @staticmethod
    def get_memory_usage() -> float:
        """Get current memory usage in MB - simplified version"""
        # Memory monitoring removed to avoid external dependencies
        # This method is kept for API compatibility but returns 0
        return 0.0
    
    @staticmethod
    def log_memory_usage(logger, operation: str):
        """Log operation completion (memory monitoring simplified)"""
        # Simply log that the operation completed
        logger.info(f"Completed {operation}")

# Performance monitoring decorator
def monitor_performance(func):
    """Decorator to monitor function performance"""
    def wrapper(*args, **kwargs):
        import time
        start_time = time.time()
        try:
            result = func(*args, **kwargs)
            duration = time.time() - start_time
            if hasattr(args[0], 'logger'):
                args[0].logger.info(f"{func.__name__} completed in {duration:.2f} seconds")
            return result
        except Exception as e:
            duration = time.time() - start_time
            if hasattr(args[0], 'logger'):
                args[0].logger.error(f"{func.__name__} failed after {duration:.2f} seconds: {e}")
            raise
    return wrapper

if __name__ == "__main__":
    root = ThemedTk(theme="arc")
    app = ConsolidatorApp(root)
    root.mainloop()